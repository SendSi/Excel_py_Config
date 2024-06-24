import openpyxl
import re

def get_end_row(sheet_obj):
    for row in range(sheet_obj.max_row, 0, -1):
        cell_obj = sheet_obj.cell(row = row, column = 1)
        if cell_obj.value == "END":
            return row

def get_column(sheet_obj, header_row):
    for column in range(sheet_obj.max_column, 0, -1):
        cell_obj = sheet_obj.cell(row = header_row, column = column)
        if cell_obj.value != None and cell_obj.value.strip() != "":
            return column

def get_header_indices(sheet_obj, header_row, start_column):
    column_indices = []
    max_column = get_column(sheet_obj, header_row)
    for i in range(start_column, max_column + 1):
        cell_obj = sheet_obj.cell(row = header_row, column = i)
        if cell_obj.value != None and cell_obj.value.strip() != "":
            column_indices.append(i)
    return column_indices

def get_row_data(sheet_obj, row, column_indices):
    row_data_list = []
    for column in column_indices:
        cell_obj = sheet_obj.cell(row, column)
        row_data_list.append(cell_obj.value)
    return row_data_list

def get_row_comments(sheet_obj, row, column_indices):
    comment_list = []
    for column in column_indices:
        cell_obj = sheet_obj.cell(row, column)
        if cell_obj.comment is None:
            comment_list.append("")
        else:
            comment_list.append(cell_obj.comment.text)
    return comment_list

def get_records(sheet_obj, start_row, end_row, column_indices):
    records = []
    for row in range(start_row, end_row + 1):
        records.append(get_row_data(sheet_obj, row, column_indices))
    return records

def get_sql_types(type_list):
    result = []
    for t in type_list:
        sqlType = None
        if t == "int":
            sqlType = "int(11) DEFAULT '0'"
        elif t == "bigint":
            sqlType = "bigint(20) DEFAULT '0'"
        elif t == "bit":
            sqlType = "bit(1) DEFAULT b'0'"
        elif t == "double":
            sqlType = "double"
        elif t == "varchar":
            sqlType = "varchar(256) DEFAULT ''"
        elif t == "longtext":
            sqlType = "longtext"
        elif t == "datetime":
            sqlType = "datetime DEFAULT '1970-01-01 00:00:00'"
        else:
            raise ValueError("field type error " + str(t))
        result.append(sqlType)
    return result

def get_str_value_by_type(value, value_type):
    if value_type == "int" or value_type == "bigint" or value_type == "double":
        return str(value) if value != None else '0'
    if value_type == "varchar" or value_type == "longtext" or value_type == "datetime":
        return f"'{value}'" if value != None else "''"
    if value_type == "bit":
        return "False" if value is None or value == 0 or value == "" or value == "0" else "True"
    else:
        raise ValueError("field type error " + str(value_type))

def get_str_value_by_type_for_lua(value, value_type):
    value_type = value_type.strip(' \r\n')
    if value_type == "int":
        if value is None:
            return '0'
        if not re.match(r"^-?\d*(.\d+)?$", str(value)):
            raise ValueError(f"{value} is not number")
        return "{:.9f}".format(float(value) + 0.00000000001).rstrip('0').rstrip('.')
    if value_type == "double" or value_type == "float":
        if value is None:
            return '0'
        return "{:.9f}".format(float(value) + 0.00000000001).rstrip('0').rstrip('.')
    if value_type == "datetime" or value_type == "bit" or value_type == "bigint" or value_type == "varchar" or value_type == "longtext":
        if value is None:
            return '""'
        if re.match(r"[\[\]\"]", str(value)):
            raise ValueError(f"{value}存在禁用字符")
        return  f'"{value}"'.replace("\r", "").replace("\n", "")
    raise ValueError("field type error " + str(value_type))

def is_not_valid_table(sheet_obj):
    a2 = sheet_obj.cell(2, 1).value
    if a2 is None or str.upper(str(a2)) != "CLIENT":
        return True
    a3 = sheet_obj.cell(3, 1).value
    if a3 is None or str.upper(str(a3)) != "TYPE":
        return True
    a4 = sheet_obj.cell(4, 1).value
    if a4 is None or str.upper(str(a4)) != "SERVER":
        return True
    end_row = get_end_row(sheet_obj)
    if end_row is None:
        return True
    return False