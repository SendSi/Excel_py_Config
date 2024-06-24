#繁体转简体
import openpyxl
import glob
from langconv import *
from os import path
from win32com.client import Dispatch

folder = sys.argv[1]
print("start python 繁体转简体")
if path.isfile(folder):
    files = [folder]
elif path.isdir(folder):
    files = [file for file in glob.glob(folder + "/**/*", recursive=True) if re.match(r"[^~]+\.xls[xm]?$", file)]

for filePath in files:
    print("start ", filePath)

    if filePath.endswith('.xlsm'):
        workbook_object = openpyxl.load_workbook(filePath, keep_vba=True)
    else:
        workbook_object = openpyxl.load_workbook(filePath)
    sheet_obj = workbook_object.active

    colCount = sheet_obj.max_column
    for index, row in enumerate(sheet_obj.rows):
        for i in range(0, colCount - 1):
            if row[i].value is None:
                continue

            if not isinstance(row[i].value, str):
                continue
            
            row[i].value = Converter('zh-hans').convert(row[i].value)
            #print('转换文字 （', index, ', ', i, ': ', row[i].value)  zh-hans简体中文       zh-hant繁体中文
    
    workbook_object.save(filename=filePath)
    
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filePath)
    xlBook.Save()
    xlBook.Close()
    
print('转换完成')
